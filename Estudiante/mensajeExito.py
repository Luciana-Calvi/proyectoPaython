import tkinter as tk
from openpyxl import load_workbook, Workbook

def guardar_datos(nombre, edad, mail, telefono, domicilio, curso):
    archivo_excel = "inscripciones.xlsx"
    try:
        excel = load_workbook(archivo_excel)
        hoja = excel.active
    except FileNotFoundError:
        excel = Workbook()
        hoja = excel.active
        hoja.append(["Nombre", "Edad", "Email", "Teléfono", "Domicilio", "Curso"])

    hoja.append([nombre, edad, mail, telefono, domicilio, curso])
    excel.save(archivo_excel)

def mensaje_exito(ventana_principal, nombre, edad, mail, telefono, domicilio, curso):
    modal = tk.Toplevel(ventana_principal)
    modal.title("Éxito")
    modal.geometry("300x150")
    modal.configure(bg="lavenderblush")
    modal.grab_set()

    tk.Label(modal, text="Formulario enviado con éxito.",
             font=("Arial", 12, "bold"),
             bg="lavenderblush", fg="darkgreen").pack(pady=15)

    def cerrar_todo():
        guardar_datos(nombre, edad, mail, telefono, domicilio, curso)
        ventana_principal.quit()

    def volver():
        guardar_datos(nombre, edad, mail, telefono, domicilio, curso)
        if ventana_principal:
            ventana_principal.deiconify()  #  Volver a mostrar la ventana principal
        modal.destroy()

    tk.Button(modal, text="Aceptar y salir", command=cerrar_todo,
              bg="palegreen", fg="black", activebackground="lightgreen").pack(pady=5)

    tk.Button(modal, text="Volver", command=volver,
              bg="lightcoral", fg="white", activebackground="salmon").pack(pady=5)



